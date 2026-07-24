import openpyxl
from openpyxl.styles import (Font, PatternFill, Alignment, Border, Side,
                              GradientFill)
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Table S1"

# ── Colour palette ──────────────────────────────────────────────────────────
COL_HEADER_BG  = "2B5240"   # dark green  – column headers
COL_HEADER_FG  = "FFFFFF"
COL_CAT_BG     = "C8DCCE"   # medium green – category rows
COL_CAT_FG     = "1A3028"
COL_SRC_BG     = "E8F2EB"   # pale green  – source sub-headers
COL_SRC_FG     = "2B5240"
COL_STRIPE     = "F3F8F4"   # very light  – alternating data rows
COL_WHITE      = "FFFFFF"
COL_SPEC_FG    = "2B5240"   # spectral type label
COL_TEX_FG     = "5A6B60"   # texture type label

def fill(hex_col):
    return PatternFill("solid", fgColor=hex_col)

def border(style="thin", color="C0CFC4"):
    s = Side(style=style, color=color)
    return Border(left=s, right=s, top=s, bottom=s)

def border_bottom(color="9DB8A5"):
    s_thin = Side(style="thin", color="D0DDD4")
    s_med  = Side(style="medium", color=color)
    return Border(left=s_thin, right=s_thin, top=s_thin, bottom=s_med)

# ── Column widths ───────────────────────────────────────────────────────────
col_widths = [4, 26, 14, 14, 16, 38]
cols = ["#", "CSV column name", "Source", "Statistic", "Type", "Description"]
for i, w in enumerate(col_widths, 1):
    ws.column_dimensions[get_column_letter(i)].width = w

# ── Helper: write a row with style ──────────────────────────────────────────
def write_row(ws, row_idx, values, bg, fg, bold=False, italic=False,
              font_size=10, align="left", border_fn=None):
    for col_idx, val in enumerate(values, 1):
        cell = ws.cell(row=row_idx, column=col_idx, value=val)
        cell.fill = fill(bg)
        cell.font = Font(name="Calibri", size=font_size, color=fg,
                         bold=bold, italic=italic)
        cell.alignment = Alignment(horizontal=align, vertical="center",
                                   wrap_text=(col_idx == 6))
        cell.border = (border_fn() if border_fn else border())
    ws.row_dimensions[row_idx].height = 16

# ── Title row (row 1) ────────────────────────────────────────────────────────
ws.merge_cells("A1:F1")
title_cell = ws["A1"]
title_cell.value = ("Table S1  |  UAV-derived covariates used in "
                    "TabICL-based weed density models")
title_cell.font  = Font(name="Calibri", size=12, bold=True, color="1A2420")
title_cell.fill  = fill("EBF3ED")
title_cell.alignment = Alignment(horizontal="left", vertical="center")
title_cell.border = border(color="C0CFC4")
ws.row_dimensions[1].height = 22

# ── Subtitle / caption (row 2) ───────────────────────────────────────────────
ws.merge_cells("A2:F2")
sub_cell = ws["A2"]
sub_cell.value = ("All 40 covariates extracted per 1 m × 1 m ROI from "
                  "multispectral UAV orthomosaics. Feature names match "
                  "column headers in ROI_features_stacked.csv.")
sub_cell.font  = Font(name="Calibri", size=9, italic=True, color="4A6358")
sub_cell.fill  = fill("EBF3ED")
sub_cell.alignment = Alignment(horizontal="left", vertical="center",
                                wrap_text=True)
sub_cell.border = border(color="C0CFC4")
ws.row_dimensions[2].height = 28

# ── Column header row (row 3) ────────────────────────────────────────────────
write_row(ws, 3, cols, COL_HEADER_BG, COL_HEADER_FG,
          bold=True, font_size=10, border_fn=border_bottom)
ws.row_dimensions[3].height = 18

# ── Data ─────────────────────────────────────────────────────────────────────
rows = [
    # (kind, values...)
    # kind: "cat"=category header, "src"=source sub-header, "data"=data row
    ("cat", "Reflectance bands", "", "", "", ""),

    ("src", "", "Green — 560 nm", "", "", ""),
    ("data", 1,  "green_mean",         "Green",     "Mean",        "Spectral", "ROI mean reflectance"),
    ("data", 2,  "green_std",          "Green",     "SD",          "Spectral", "ROI standard deviation"),
    ("data", 3,  "green_contrast",     "Green",     "Contrast",    "GLCM texture", "Local intensity variation"),
    ("data", 4,  "green_entropy",      "Green",     "Entropy",     "GLCM texture", "Disorder of gray-level distribution"),
    ("data", 5,  "green_homogeneity",  "Green",     "Homogeneity", "GLCM texture", "Closeness of distribution to diagonal"),

    ("src", "", "Red — 650 nm", "", "", ""),
    ("data", 6,  "red_mean",           "Red",       "Mean",        "Spectral", "ROI mean reflectance"),
    ("data", 7,  "red_std",            "Red",       "SD",          "Spectral", "ROI standard deviation"),
    ("data", 8,  "red_contrast",       "Red",       "Contrast",    "GLCM texture", "Local intensity variation"),
    ("data", 9,  "red_entropy",        "Red",       "Entropy",     "GLCM texture", "Disorder of gray-level distribution"),
    ("data", 10, "red_homogeneity",    "Red",       "Homogeneity", "GLCM texture", "Closeness of distribution to diagonal"),

    ("src", "", "Red-edge — 730 nm", "", "", ""),
    ("data", 11, "red_edge_mean",        "Red-edge",  "Mean",        "Spectral", "ROI mean reflectance"),
    ("data", 12, "red_edge_std",         "Red-edge",  "SD",          "Spectral", "ROI standard deviation"),
    ("data", 13, "red_edge_contrast",    "Red-edge",  "Contrast",    "GLCM texture", "Local intensity variation"),
    ("data", 14, "red_edge_entropy",     "Red-edge",  "Entropy",     "GLCM texture", "Disorder of gray-level distribution"),
    ("data", 15, "red_edge_homogeneity", "Red-edge",  "Homogeneity", "GLCM texture", "Closeness of distribution to diagonal"),

    ("src", "", "Near-infrared — 840 nm", "", "", ""),
    ("data", 16, "nir_mean",           "NIR",       "Mean",        "Spectral", "ROI mean reflectance"),
    ("data", 17, "nir_std",            "NIR",       "SD",          "Spectral", "ROI standard deviation"),
    ("data", 18, "nir_contrast",       "NIR",       "Contrast",    "GLCM texture", "Local intensity variation"),
    ("data", 19, "nir_entropy",        "NIR",       "Entropy",     "GLCM texture", "Disorder of gray-level distribution"),
    ("data", 20, "nir_homogeneity",    "NIR",       "Homogeneity", "GLCM texture", "Closeness of distribution to diagonal"),

    ("cat", "Vegetation indices", "", "", "", ""),

    ("src", "", "NDVI — Normalized Difference Vegetation Index  [(NIR − Red) / (NIR + Red)]", "", "", ""),
    ("data", 21, "NDVI_mean",          "NDVI",      "Mean",        "Spectral", "ROI mean index value"),
    ("data", 22, "NDVI_std",           "NDVI",      "SD",          "Spectral", "ROI standard deviation"),
    ("data", 23, "NDVI_contrast",      "NDVI",      "Contrast",    "GLCM texture", "Local intensity variation"),
    ("data", 24, "NDVI_entropy",       "NDVI",      "Entropy",     "GLCM texture", "Disorder of gray-level distribution"),
    ("data", 25, "NDVI_homogeneity",   "NDVI",      "Homogeneity", "GLCM texture", "Closeness of distribution to diagonal"),

    ("src", "", "NDRE — Normalized Difference Red-Edge Index  [(NIR − Red-edge) / (NIR + Red-edge)]", "", "", ""),
    ("data", 26, "NDRE_mean",          "NDRE",      "Mean",        "Spectral", "ROI mean index value"),
    ("data", 27, "NDRE_std",           "NDRE",      "SD",          "Spectral", "ROI standard deviation"),
    ("data", 28, "NDRE_contrast",      "NDRE",      "Contrast",    "GLCM texture", "Local intensity variation"),
    ("data", 29, "NDRE_entropy",       "NDRE",      "Entropy",     "GLCM texture", "Disorder of gray-level distribution"),
    ("data", 30, "NDRE_homogeneity",   "NDRE",      "Homogeneity", "GLCM texture", "Closeness of distribution to diagonal"),

    ("src", "", "CIgreen — Chlorophyll Index Green  [(NIR / Green) − 1]", "", "", ""),
    ("data", 31, "CIgreen_mean",         "CIgreen",   "Mean",        "Spectral", "ROI mean index value"),
    ("data", 32, "CIgreen_std",          "CIgreen",   "SD",          "Spectral", "ROI standard deviation"),
    ("data", 33, "CIgreen_contrast",     "CIgreen",   "Contrast",    "GLCM texture", "Local intensity variation"),
    ("data", 34, "CIgreen_entropy",      "CIgreen",   "Entropy",     "GLCM texture", "Disorder of gray-level distribution"),
    ("data", 35, "CIgreen_homogeneity",  "CIgreen",   "Homogeneity", "GLCM texture", "Closeness of distribution to diagonal"),

    ("src", "", "CIred-edge — Chlorophyll Index Red-Edge  [(NIR / Red-edge) − 1]", "", "", ""),
    ("data", 36, "CIrededge_mean",        "CIred-edge", "Mean",        "Spectral", "ROI mean index value"),
    ("data", 37, "CIrededge_std",         "CIred-edge", "SD",          "Spectral", "ROI standard deviation"),
    ("data", 38, "CIrededge_contrast",    "CIred-edge", "Contrast",    "GLCM texture", "Local intensity variation"),
    ("data", 39, "CIrededge_entropy",     "CIred-edge", "Entropy",     "GLCM texture", "Disorder of gray-level distribution"),
    ("data", 40, "CIrededge_homogeneity", "CIred-edge", "Homogeneity", "GLCM texture", "Closeness of distribution to diagonal"),
]

data_row_counter = 0
excel_row = 4

for entry in rows:
    kind = entry[0]

    if kind == "cat":
        ws.merge_cells(f"A{excel_row}:F{excel_row}")
        cell = ws.cell(row=excel_row, column=1, value=entry[1])
        cell.fill  = fill(COL_CAT_BG)
        cell.font  = Font(name="Calibri", size=10, bold=True, color=COL_CAT_FG)
        cell.alignment = Alignment(horizontal="left", vertical="center",
                                   indent=1)
        cell.border = border_bottom(color="7AAD8E")
        for c in range(2, 7):
            ws.cell(row=excel_row, column=c).fill = fill(COL_CAT_BG)
            ws.cell(row=excel_row, column=c).border = border_bottom(color="7AAD8E")
        ws.row_dimensions[excel_row].height = 17
        data_row_counter = 0  # reset stripe counter

    elif kind == "src":
        ws.merge_cells(f"B{excel_row}:F{excel_row}")
        # col A blank
        ws.cell(row=excel_row, column=1).fill = fill(COL_SRC_BG)
        ws.cell(row=excel_row, column=1).border = border(color="C8D9CC")
        cell = ws.cell(row=excel_row, column=2, value=entry[2])
        cell.fill  = fill(COL_SRC_BG)
        cell.font  = Font(name="Calibri", size=9, bold=True, italic=True,
                          color=COL_SRC_FG)
        cell.alignment = Alignment(horizontal="left", vertical="center",
                                   indent=1)
        cell.border = border(color="C8D9CC")
        for c in range(3, 7):
            ws.cell(row=excel_row, column=c).fill = fill(COL_SRC_BG)
            ws.cell(row=excel_row, column=c).border = border(color="C8D9CC")
        ws.row_dimensions[excel_row].height = 15

    elif kind == "data":
        _, num, csv_name, source, stat, typ, desc = entry
        bg = COL_STRIPE if data_row_counter % 2 == 1 else COL_WHITE
        data_row_counter += 1

        vals = [num, csv_name, source, stat, typ, desc]
        for col_idx, val in enumerate(vals, 1):
            cell = ws.cell(row=excel_row, column=col_idx, value=val)
            cell.fill   = fill(bg)
            cell.border = border(color="D4E0D8")
            cell.alignment = Alignment(horizontal="center" if col_idx == 1 else "left",
                                       vertical="center", wrap_text=(col_idx == 6))
            # type column colour
            if col_idx == 5:
                fg = COL_SPEC_FG if "Spectral" in str(val) else COL_TEX_FG
                bold = "Spectral" in str(val)
            else:
                fg  = "1A2420"
                bold = False
            # CSV name in monospace-ish style
            if col_idx == 2:
                cell.font = Font(name="Courier New", size=9, color="1A3028")
            else:
                cell.font = Font(name="Calibri", size=10, color=fg, bold=bold)
        ws.row_dimensions[excel_row].height = 15

    excel_row += 1

# ── GLCM footnote ────────────────────────────────────────────────────────────
excel_row += 1
ws.merge_cells(f"A{excel_row}:F{excel_row}")
note1 = ws.cell(row=excel_row, column=1,
    value=("GLCM parameters: ROI pixel values quantized to 32 gray levels; "
           "co-occurrence matrices computed with a one-pixel offset at four "
           "orientations (0°, 45°, 90°, 135°); directional estimates averaged "
           "to obtain rotationally invariant descriptors (Haralick et al., 1973)."))
note1.font      = Font(name="Calibri", size=9, italic=True, color="4A6358")
note1.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
note1.fill      = fill("EBF3ED")
note1.border    = border(color="C8D9CC")
ws.row_dimensions[excel_row].height = 28

excel_row += 1
ws.merge_cells(f"A{excel_row}:F{excel_row}")
note2 = ws.cell(row=excel_row, column=1,
    value=("Statistic definitions — "
           "Contrast = Σ(i−j)²·p(i,j);  "
           "Entropy = −Σ p(i,j)·log₂ p(i,j);  "
           "Homogeneity = Σ p(i,j) / (1 + |i−j|),  "
           "where p(i,j) is the normalised co-occurrence frequency for gray levels i and j."))
note2.font      = Font(name="Calibri", size=9, italic=True, color="4A6358")
note2.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
note2.fill      = fill("EBF3ED")
note2.border    = border(color="C8D9CC")
ws.row_dimensions[excel_row].height = 28

# ── Freeze panes below header ─────────────────────────────────────────────────
ws.freeze_panes = "A4"

# ── Save ──────────────────────────────────────────────────────────────────────
out_path = "figures/Table_S1_covariates.xlsx"
wb.save(out_path)
print(f"Saved: {out_path}")
